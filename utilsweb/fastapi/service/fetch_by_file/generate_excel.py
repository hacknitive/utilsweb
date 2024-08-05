from datetime import datetime
from io import BytesIO

from fastapi import status
from utilscommon import make_flat
from openpyxl import Workbook

MEDIA_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def generate_excel(
        file_name: str,
        data: dict,
) -> dict:
    now = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    file_name = f"{file_name}-Result-{now}.xlsx"

    wb = Workbook()

    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    for sheet_name, data_i in data.items():
        wb = create_sheet(
            data=data_i,
            sheet_name=sheet_name,
            wb=wb,
        )

    file_in_bytesio = BytesIO()
    wb.save(file_in_bytesio)
    file_in_bytesio.seek(0)

    return {
        'media_type': MEDIA_TYPE,
        'content': file_in_bytesio,
        'status_code': status.HTTP_200_OK,
        'headers': {"Content-Disposition": f"attachment; filename={file_name}"},
    }


def create_sheet(
        data: dict | list,
        sheet_name: str,
        wb: Workbook
):
    if isinstance(data, list):
        flat_data = [make_flat(i) for i in data]

    else:
        flat_data = [make_flat(
            data=data,
        )]

    headers = set()
    for i in flat_data:
        headers.update(set(i.keys()))
    headers = list(headers)
    headers.sort()

    ws1 = wb.create_sheet(title=sheet_name.title())
    ws1.append(headers)
    for row in flat_data:
        ws1.append([row.get(header, '') for header in headers])

    return wb
